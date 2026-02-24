# frontend/app.py

import streamlit as st
import pandas as pd
import sys
import os
import io
from datetime import datetime

# Add backend path
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "backend")))
from classifier import predict_ticket

st.set_page_config(page_title="Ticket Intelligence Dashboard", layout="wide")
st.title("📊 Ticket Intelligence & Categorization System")

uploaded_file = st.file_uploader("Upload Ticket Excel File", type=["xlsx"])

if uploaded_file:

    df = pd.read_excel(uploaded_file)
    df.columns = df.columns.str.strip()

    predictions = []
    confidences = []

    # ===============================
    # CLEAN INPUT TEXT (IMPORTANT)
    # ===============================
    for _, row in df.iterrows():

        text = " ".join([
            str(row.get("Ticket Summary", "")),
            str(row.get("Ticket Description", ""))
        ])

        result = predict_ticket(text)

        predictions.append(result["category"])
        confidences.append(result["confidence"])

    df["Predicted Category"] = predictions
    df["Confidence"] = confidences

    # ===============================
    # PROFESSIONAL EXECUTIVE SUMMARY
    # ===============================
    def refine_summary(row):

        summary = str(row.get("Ticket Summary", "")).strip()
        description = str(row.get("Ticket Description", "")).strip()
        category = str(row.get("Predicted Category", "")).strip()

        combined = f"{summary} {description}"
        combined = " ".join(combined.split())[:220]

        if not combined:
            combined = "Ticket details insufficient for executive analysis."

        return f"{category} identified impacting operational workflow. {combined}"

    df["Executive Summary"] = df.apply(refine_summary, axis=1)

    # ===============================
    # MONTH HANDLING
    # ===============================
    date_col = None
    for possible in ["Closed On", "Reported On", "Resolved on", "Raised on"]:
        if possible in df.columns:
            date_col = possible
            break

    if date_col:
        df["Date"] = pd.to_datetime(df[date_col], errors="coerce", dayfirst=True)
        df["Month"] = df["Date"].dt.strftime("%B")
    else:
        df["Month"] = "Unknown"

    # ===============================
    # KPI SECTION
    # ===============================
    st.markdown("## 📌 Key Performance Indicators")

    col1, col2, col3, col4 = st.columns(4)

    total_tickets = len(df)
    unique_categories = df["Predicted Category"].nunique()
    avg_confidence = round(df["Confidence"].mean(), 3)
    top_category = df["Predicted Category"].value_counts().idxmax()

    col1.metric("Total Tickets", total_tickets)
    col2.metric("Unique Categories", unique_categories)
    col3.metric("Avg Confidence", avg_confidence)
    col4.metric("Top Category", top_category)

    # ===============================
    # DASHBOARD VISUALS
    # ===============================
    st.markdown("## 📊 Dashboard Overview")

    st.subheader("Issue Category Distribution")
    st.bar_chart(df["Predicted Category"].value_counts())

    st.subheader("Month-wise Ticket Count")
    st.bar_chart(df.groupby("Month").size())

    st.subheader("Month-wise Issue Percentage")

    month_category = (
        df.groupby(["Month", "Predicted Category"])
          .size()
          .unstack(fill_value=0)
    )

    month_percentage = (
        month_category.div(month_category.sum(axis=1), axis=0) * 100
    )

    st.dataframe(month_percentage.round(2))
    st.bar_chart(month_percentage)

    # ===============================
    # DETAILED TABLE
    # ===============================
    st.markdown("## 📄 Detailed Ticket View")
    st.dataframe(df)

    # ===============================
    # PROFESSIONAL EXCEL EXPORT
    # ===============================
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:

        df.to_excel(writer, sheet_name="Detailed_Tickets", index=False)
        workbook = writer.book

        # ----------------------------
        # CATEGORY SUMMARY
        # ----------------------------
        category_summary = (
            df["Predicted Category"]
            .value_counts()
            .reset_index()
        )
        category_summary.columns = ["Category", "Total Tickets"]

        category_summary["Percentage"] = (
            category_summary["Total Tickets"] /
            category_summary["Total Tickets"].sum() * 100
        ).round(2)

        category_summary.to_excel(writer, sheet_name="Category_Summary", index=False)

        ws_summary = writer.sheets["Category_Summary"]

        from openpyxl.styles import Font, PatternFill
        header_fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")

        for cell in ws_summary[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill

        # ----------------------------
        # CATEGORY BAR CHART
        # ----------------------------
        from openpyxl.chart import BarChart, Reference

        chart = BarChart()
        chart.title = "Issue Category Distribution"
        chart.y_axis.title = "Ticket Count"
        chart.x_axis.title = "Category"
        chart.style = 10

        data = Reference(ws_summary,
                         min_col=2,
                         min_row=1,
                         max_row=len(category_summary)+1)

        categories = Reference(ws_summary,
                               min_col=1,
                               min_row=2,
                               max_row=len(category_summary)+1)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.width = 20
        chart.height = 12

        ws_summary.add_chart(chart, "E2")

        # ----------------------------
        # MONTHLY TREND
        # ----------------------------
        monthly_pivot = (
            df.groupby(["Month", "Predicted Category"])
            .size()
            .unstack(fill_value=0)
        )

        monthly_pivot.to_excel(writer, sheet_name="Monthly_Trend")

        ws_month = writer.sheets["Monthly_Trend"]

        for cell in ws_month[1]:
            cell.font = Font(bold=True)



        # ==========================================
        # AREA-WISE PIVOT TABLE (SAFE VERSION)
        # ==========================================
        
        area_column = None
        
        # Detect Area column dynamically
        for col in df.columns:
            if col.strip().lower() in ["area", "Track", "area category", "track"]:
                area_column = col
                break
        
        if area_column:
        
            area_pivot = pd.pivot_table(
                df,
                index=[area_column, "Predicted Category"],
                columns="Month",
                aggfunc="size",   # <-- IMPORTANT FIX
                fill_value=0
            )
        
            # Add Grand Total column
            area_pivot["Grand Total"] = area_pivot.sum(axis=1)
        
            area_pivot = area_pivot.reset_index()
        
            area_pivot.to_excel(writer, sheet_name="Area_Wise_Pivot", index=False)
        
            ws_area = writer.sheets["Area_Wise_Pivot"]
        
            from openpyxl.styles import Font, PatternFill
        
            header_fill = PatternFill(start_color="1F4E78",
                                      end_color="1F4E78",
                                      fill_type="solid")
        
            for cell in ws_area[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = header_fill
        
        else:
            pd.DataFrame({
                "Message": ["Area column not available in uploaded file"]
            }).to_excel(writer, sheet_name="Area_Wise_Pivot", index=False)
        
       
        # ----------------------------
        # STACKED MONTHLY CHART
        # ----------------------------
        stacked_chart = BarChart()
        stacked_chart.type = "col"
        stacked_chart.grouping = "stacked"
        stacked_chart.title = "Monthly Issue Distribution"
        stacked_chart.y_axis.title = "Tickets"
        stacked_chart.x_axis.title = "Month"

        data = Reference(ws_month,
                         min_col=2,
                         min_row=1,
                         max_row=ws_month.max_row,
                         max_col=ws_month.max_column)

        categories = Reference(ws_month,
                               min_col=1,
                               min_row=2,
                               max_row=ws_month.max_row)

        stacked_chart.add_data(data, titles_from_data=True)
        stacked_chart.set_categories(categories)

        stacked_chart.width = 22
        stacked_chart.height = 12

        ws_month.add_chart(stacked_chart, "E2")

    output.seek(0)

    st.download_button(
        label="📥 Download Professional Executive Report",
        data=output,
        file_name=f"Ticket_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
